# ============================================================
# HAY STRAW LENGTH MEASUREMENT TOOL  —  MERGED v3.0
# ============================================================
# Combines:
#   • Script A  — clean BFS skeleton measurement, overlap filter,
#                 component cleanup, length guard-rails
#   • Script B  — full pipeline: dual HSV/YOLO modes, auto-
#                 calibration from a quarter coin, Dijkstra-based
#                 overlap reconstruction, Excel + CSV export with
#                 embedded overlay images, structured logging
#
# QUICK-START
# -----------
#   1. Set YOLO_MODEL_PATH to your .pt file.
#   2. Run.  A file-picker opens (or hard-code paths in DEBUG_FILES).
#   3. Results appear in straw_results.xlsx / straw_results.csv.
#
# MODES
# -----
#   DETECTION_MODE = "yolo"  →  YOLOv11 instance segmentation
#   DETECTION_MODE = "hsv"   →  Classic HSV color threshold
#
# CALIBRATION
# -----------
#   CALIBRATION_MODE = "auto"   →  Quarter coin detected by YOLO
#   CALIBRATION_MODE = "manual" →  You enter mm/px at runtime
#   CALIBRATION_PER_IMAGE = True  →  Re-calibrate every image
#   CALIBRATION_PER_IMAGE = False →  Calibrate once, reuse
#
# MEASUREMENT STRATEGY
# --------------------
#   No overlap  → BFS longest-path  (fast, accurate)
#   Overlap     → Dijkstra between paired endpoints (reconstructs
#                 individual straws through crossings)
# ============================================================

import cv2
import numpy as np
import os
import sys
import time
import heapq
import logging
import pandas as pd
from collections import deque
from itertools import combinations
from skimage.morphology import skeletonize
from scipy.spatial.distance import cdist
from scipy import ndimage as ndi
from tkinter import Tk, filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ============================================================
# ❶  CONFIGURATION  — edit everything here
# ============================================================

# --- Model & detection ---
DETECTION_MODE   = "yolo"           # "yolo" or "hsv"
YOLO_MODEL_PATH  = "best_K_Fold.pt" # path to your trained .pt
YOLO_CONF        = 0.5              # detection confidence threshold

# --- Calibration ---
CALIBRATION_MODE      = "auto"      # "auto" or "manual"
CALIBRATION_PER_IMAGE = True        # True = recalibrate each image
QUARTER_CLASS_NAME    = "quarter"   # must match your YOLO class label
QUARTER_DIAMETER_MM   = 24.26       # US quarter diameter (mm)
PIXEL_TO_MM           = 0.55        # fallback / manual default (mm/px)

# --- Length guard-rails (applied after scaling) ---
MIN_LENGTH_MM = 10
MAX_LENGTH_MM = 1500    # raised — long straws can exceed 400 mm

# --- Skeleton / mask filters ---
MIN_PATH_LENGTH_PX = 10   # reject skeletons shorter than this (px)
MIN_MASK_AREA_PX   = 200  # reject YOLO masks smaller than this (px²)
MIN_SIZE_COMPONENT = 200  # min connected-component size after cleanup
MIN_CONTOUR_AREA   = 10   # min HSV contour area (px²)
IOU_THRESH         = 0.7  # overlap filter: discard lower-conf duplicate

# --- Overlap reconstruction ---
ANGLE_TOLERANCE_DEG = 35.0  # how close to 180° for a valid endpoint pair
CLUSTER_RADIUS_PX   = 80    # search radius around branch-point centroid

# --- Shape classification ---
CURVE_THRESHOLD = 1.1   # path/euclidean ratio above which → "Curved"

# --- HSV range for straw colour (only used in HSV mode) ---
STRAW_HSV_LOWER = np.array([10,  40,  80])
STRAW_HSV_UPPER = np.array([35, 255, 255])

# --- Debug ---
DEBUG_MODE = True           # show popup windows in HSV mode
LOG_FILE   = "straw_debug.log"

# --- Hard-code image paths here to skip the file picker ---
DEBUG_FILES = [
    # r"C:\Users\Logan\Documents\...\IMG_2733.jpg",
]

# ============================================================
# ❷  LOGGING
# ============================================================

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, mode="w"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# Runtime flag — suppresses cv2.imshow in YOLO mode
_SUPPRESS_POPUPS = False


def debug_step(name, image, wait=200):
    """Show a debug window unless popups are suppressed."""
    if _SUPPRESS_POPUPS or not DEBUG_MODE or image is None:
        return
    cv2.imshow(f"DEBUG: {name}", image)
    cv2.waitKey(wait)

# ============================================================
# ❸  FILE PICKER
# ============================================================

def select_files():
    """Return list of image paths from DEBUG_FILES or a GUI picker."""
    if DEBUG_FILES:
        valid   = [f for f in DEBUG_FILES if os.path.exists(f)]
        missing = [f for f in DEBUG_FILES if not os.path.exists(f)]
        for m in missing:
            log.warning(f"DEBUG_FILES: not found — {m}")
        log.info(f"DEBUG_FILES mode: {len(valid)} file(s)")
        return valid

    log.info("Opening file picker…")
    try:
        root = Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        files = list(filedialog.askopenfilenames(
            filetypes=[("Image files", "*.jpg *.jpeg *.png")]
        ))
        root.destroy()
    except Exception as exc:
        log.error(f"File picker failed: {exc}")
        log.error("Add image paths to DEBUG_FILES in the config section.")
        return []
    log.info(f"Selected {len(files)} file(s)")
    return files

# ============================================================
# ❹  YOLO LOADER
# ============================================================

def load_yolo_model(path):
    try:
        from ultralytics import YOLO
        if not os.path.exists(path):
            log.warning(f"YOLO model not found: '{path}'. Falling back to HSV.")
            return None
        model = YOLO(path)
        log.info(f"YOLO model loaded: {path}")
        log.info(f"  Classes: {model.names}")
        return model
    except ImportError:
        log.warning("ultralytics not installed. Run: pip install ultralytics")
        return None
    except Exception as exc:
        log.warning(f"YOLO load failed: {exc}. Falling back to HSV.")
        return None

# ============================================================
# ❺  OVERLAP FILTER  (keep highest-confidence non-overlapping masks)
# ============================================================

def filter_overlapping_masks(masks, iou_thresh=IOU_THRESH):
    """
    Given a list of mask dicts (each with 'mask' and 'confidence'),
    discard lower-confidence masks whose IoU with a kept mask exceeds
    iou_thresh.  Returns the filtered list.
    """
    kept = []
    for m in sorted(masks, key=lambda x: -x["confidence"]):
        m_bin = m["mask"] > 0
        keep  = True
        for k in kept:
            k_bin = k["mask"] > 0
            inter = np.logical_and(m_bin, k_bin).sum()
            union = np.logical_or(m_bin, k_bin).sum()
            if union > 0 and inter / union > iou_thresh:
                keep = False
                break
        if keep:
            kept.append(m)
    return kept

# ============================================================
# ❻  MASK EXTRACTION — YOLO
# ============================================================

def get_masks_yolo(image, model):
    """
    Run YOLO inference and return one mask dict per straw detection.
    Quarter detections are skipped (used only for calibration).
    Applies the overlap filter before returning.
    """
    h_img, w_img = image.shape[:2]
    results      = model(image, conf=YOLO_CONF, verbose=False)[0]
    names        = model.names
    masks_raw    = []
    has_masks    = results.masks is not None

    log.debug(f"  Raw YOLO detections: {len(results.boxes)}")
    if not has_masks:
        log.warning("  YOLO returned no segmentation masks — "
                    "ensure best.pt is an instance-segmentation model.")

    for idx, box in enumerate(results.boxes):
        cls_id   = int(box.cls[0])
        cls_name = names.get(cls_id, "").lower().strip()
        conf     = float(box.conf[0])

        if cls_name == QUARTER_CLASS_NAME.lower().strip():
            log.debug(f"  Skipping quarter detection (idx={idx})")
            continue

        x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
        x1 = max(0, x1); y1 = max(0, y1)
        x2 = min(w_img, x2); y2 = min(h_img, y2)

        full_mask = np.zeros((h_img, w_img), dtype=np.uint8)

        if has_masks and idx < len(results.masks.data):
            mask_raw = results.masks.data[idx].cpu().numpy()
            mask_u8  = (mask_raw * 255).astype(np.uint8)
            mask_u8  = cv2.resize(mask_u8, (w_img, h_img),
                                  interpolation=cv2.INTER_NEAREST)
            _, mask_bin = cv2.threshold(mask_u8, 127, 255, cv2.THRESH_BINARY)
            kernel      = np.ones((3, 3), np.uint8)
            mask_bin    = cv2.morphologyEx(mask_bin, cv2.MORPH_CLOSE, kernel)
            full_mask   = mask_bin
        else:
            # Fallback: filled bounding box
            if (x2 - x1) >= 5 and (y2 - y1) >= 5:
                full_mask[y1:y2, x1:x2] = 255
                log.debug(f"  idx={idx}: using bbox-fill fallback")

        area = int(np.sum(full_mask > 0))
        if area < MIN_MASK_AREA_PX:
            log.debug(f"  idx={idx}: mask area {area}px < {MIN_MASK_AREA_PX}, skipping")
            continue

        masks_raw.append({
            "mask":       full_mask,
            "confidence": conf,
            "source":     "yolo",
            "class":      cls_name,
            "bbox":       (x1, y1, x2, y2),
        })

    log.debug(f"  Masks before overlap filter: {len(masks_raw)}")
    masks_filtered = filter_overlapping_masks(masks_raw)
    log.info(f"  YOLO masks after overlap filter: {len(masks_filtered)}")
    return masks_filtered, results

# ============================================================
# ❼  MASK EXTRACTION — HSV
# ============================================================

def get_masks_hsv(image, hsv_lower, hsv_upper):
    """Colour-threshold the whole image and return one mask per contour."""
    hsv    = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
    mask   = cv2.inRange(hsv, hsv_lower, hsv_upper)
    debug_step("HSV Raw Mask", mask)

    kernel = np.ones((3, 3), np.uint8)
    mask   = cv2.morphologyEx(mask, cv2.MORPH_OPEN,  kernel)
    mask   = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)
    debug_step("HSV After Morphology", mask)

    dist   = cv2.distanceTransform(mask, cv2.DIST_L2, 5)
    _, sfg = cv2.threshold(dist, 0.5 * dist.max(), 255, 0)
    sfg    = np.uint8(sfg)
    unk    = cv2.subtract(mask, sfg)
    mrk, n = ndi.label(sfg)
    log.debug(f"  HSV watershed markers: {n}")
    mrk[unk == 255] = 0
    mrk             = mrk + 1
    mrk             = cv2.watershed(cv2.cvtColor(mask, cv2.COLOR_GRAY2BGR), mrk)
    mask[mrk == -1] = 0
    debug_step("HSV After Watershed", mask)

    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL,
                                   cv2.CHAIN_APPROX_SIMPLE)
    log.info(f"  HSV contours: {len(contours)}")
    masks = []
    for cnt in contours:
        if cv2.contourArea(cnt) < MIN_CONTOUR_AREA:
            continue
        m = np.zeros(mask.shape, np.uint8)
        cv2.drawContours(m, [cnt], -1, 255, -1)
        masks.append({"mask": m, "confidence": 1.0, "source": "hsv"})
    log.info(f"  HSV valid masks: {len(masks)}")
    return masks

# ============================================================
# ❽  HSV TUNING WINDOW  (interactive sliders)
# ============================================================

def get_hsv(sample_path):
    """Open an interactive HSV tuning window for the first image."""
    log.info(f"Opening HSV tuning: {sample_path}")
    img = cv2.imread(sample_path)
    if img is None:
        log.error(f"Could not read: {sample_path}")
        sys.exit(1)
    img = cv2.resize(img, (800, 600))
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)

    WIN = "HSV Tuning  (Q to confirm)"
    cv2.namedWindow(WIN, cv2.WINDOW_NORMAL)
    cv2.resizeWindow(WIN, 800, 700)
    state = {"H Min": 0, "H Max": 179, "S Min": 0, "S Max": 255,
             "V Min": 0, "V Max": 255}

    def update(_=None):
        lower = np.array([state["H Min"], state["S Min"], state["V Min"]])
        upper = np.array([state["H Max"], state["S Max"], state["V Max"]])
        mask  = cv2.inRange(hsv, lower, upper)
        ov    = img.copy(); ov[mask == 0] = (40, 40, 40)
        hl    = img.copy(); hl[mask > 0]  = (0, 220, 80)
        bl    = cv2.addWeighted(img, 0.5, hl, 0.5, 0)
        mb    = cv2.cvtColor(mask, cv2.COLOR_GRAY2BGR)
        top   = np.hstack([img, ov])
        bot   = np.hstack([bl,  mb])
        for panel, label, pos in [
            (top, "Original",        (10, 25)),
            (top, "Masked Overlay",  (810, 25)),
            (bot, "Green Highlight", (10, 25)),
            (bot, "Binary Mask",     (810, 25)),
        ]:
            cv2.putText(panel, label, pos, cv2.FONT_HERSHEY_SIMPLEX,
                        0.7, (255, 255, 255), 2, cv2.LINE_AA)
        cv2.imshow(WIN, cv2.resize(np.vstack([top, bot]), (800, 600)))
        print(f"\r  HSV [{state['H Min']:3d}-{state['H Max']:3d}]"
              f"[{state['S Min']:3d}-{state['S Max']:3d}]"
              f"[{state['V Min']:3d}-{state['V Max']:3d}]  "
              f"px:{int(np.sum(cv2.inRange(hsv, np.array([state['H Min'], state['S Min'], state['V Min']]),\
 np.array([state['H Max'], state['S Max'], state['V Max']])) > 0)):6d}",
              end="", flush=True)

    def make_cb(key):
        def cb(val): state[key] = val; update()
        return cb

    for name, maxv, default in [
        ("H Min", 179, 0), ("H Max", 179, 179),
        ("S Min", 255, 0), ("S Max", 255, 255),
        ("V Min", 255, 0), ("V Max", 255, 255),
    ]:
        cv2.createTrackbar(name, WIN, default, maxv, make_cb(name))

    print("\n--- HSV TUNING ---  Move sliders. Q to confirm.\n")
    update()
    while True:
        if cv2.waitKey(50) & 0xFF in (ord("q"), ord("s"), 27):
            break
    cv2.destroyAllWindows()
    lower = np.array([state["H Min"], state["S Min"], state["V Min"]])
    upper = np.array([state["H Max"], state["S Max"], state["V Max"]])
    log.info(f"  HSV Lower: {lower}  Upper: {upper}")
    return lower, upper

# ============================================================
# ❾  MASK CLEANUP
# ============================================================

def clean_mask(mask):
    """Morphological close then open to fill holes and remove pepper noise."""
    kernel = np.ones((3, 3), np.uint8)
    mask   = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel, iterations=2)
    mask   = cv2.morphologyEx(mask, cv2.MORPH_OPEN,  kernel, iterations=1)
    return mask


def merge_small_components(mask, min_size=MIN_SIZE_COMPONENT):
    """Remove connected components smaller than min_size pixels."""
    num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(
        (mask > 0).astype(np.uint8), connectivity=8
    )
    cleaned = np.zeros_like(mask)
    for label in range(1, num_labels):
        if stats[label, cv2.CC_STAT_AREA] >= min_size:
            cleaned[labels == label] = 255
    return cleaned

# ============================================================
# ❿  BFS LONGEST PATH  (fast, used when no overlap)
# ============================================================

def _bfs_farthest(point_set, start):
    """BFS from start; returns (farthest_point, distance)."""
    visited  = {start}
    queue    = deque([(start, 0.0)])
    farthest = (start, 0.0)
    while queue:
        cur, dist = queue.popleft()
        if dist > farthest[1]:
            farthest = (cur, dist)
        for dx in (-1, 0, 1):
            for dy in (-1, 0, 1):
                if dx == 0 and dy == 0:
                    continue
                nb = (cur[0] + dx, cur[1] + dy)
                if nb in point_set and nb not in visited:
                    visited.add(nb)
                    queue.append((nb, dist + np.hypot(dx, dy)))
    return farthest


def longest_path_bfs(points):
    """
    Double-BFS on the skeleton point set to approximate the longest path.
    O(N) — fast enough for real-time use.
    """
    point_set = set(map(tuple, points))
    start     = next(iter(point_set))
    far1      = _bfs_farthest(point_set, start)
    far2      = _bfs_farthest(point_set, far1[0])
    return far2[1]   # length in pixels

# ============================================================
# ⓫  DIJKSTRA  (used for overlap reconstruction)
# ============================================================

def _build_graph(points):
    """Build an adjacency graph from skeleton pixels (8-connectivity)."""
    pts   = np.array(points)
    graph = {}
    for p in pts:
        tp        = tuple(p)
        graph[tp] = [
            tuple(q) for q in pts
            if 0 < np.linalg.norm(p - q) <= np.sqrt(2)
        ]
    return graph


def dijkstra_path(graph, start, end):
    """Return (distance, path_nodes) between start and end."""
    queue = [(0.0, start, [start])]
    dist  = {n: float("inf") for n in graph}
    dist[start] = 0.0
    while queue:
        d, cur, path = heapq.heappop(queue)
        if cur == end:
            return d, path
        for nb in graph.get(cur, []):
            nd = d + np.linalg.norm(np.array(cur) - np.array(nb))
            if nd < dist[nb]:
                dist[nb] = nd
                heapq.heappush(queue, (nd, nb, path + [nb]))
    return float("inf"), []

# ============================================================
# ⓬  SKELETON ANALYSIS
# ============================================================

def classify_skeleton_points(points):
    """Split skeleton pixels into endpoints (1 neighbour), body (2), branches (3+)."""
    point_set             = set(map(tuple, points))
    dirs                  = [(-1,-1),(-1,0),(-1,1),(0,-1),(0,1),(1,-1),(1,0),(1,1)]
    endpoints, body, branches = [], [], []
    for p in points:
        n  = sum(1 for d in dirs if (p[0]+d[0], p[1]+d[1]) in point_set)
        tp = tuple(p)
        if   n == 1: endpoints.append(tp)
        elif n >= 3: branches.append(tp)
        else:        body.append(tp)
    log.debug(f"  Skeleton — endpoints:{len(endpoints)}  "
              f"branches:{len(branches)}  body:{len(body)}")
    return endpoints, body, branches

# ============================================================
# ⓭  ENDPOINT CLUSTERING & PAIRING  (overlap reconstruction)
# ============================================================

def _endpoint_exit_angle(ep, graph, look_ahead=8):
    """Angle (degrees) of the direction a skeleton endpoint points."""
    path, visited, current = [ep], {ep}, ep
    for _ in range(look_ahead):
        nbs = [nb for nb in graph.get(current, []) if nb not in visited]
        if not nbs:
            break
        current = nbs[0]
        visited.add(current)
        path.append(current)
    if len(path) < 2:
        return 0.0
    dy = path[-1][0] - path[0][0]
    dx = path[-1][1] - path[0][1]
    return np.degrees(np.arctan2(dy, dx))


def cluster_and_pair_endpoints(endpoints, branch_points, graph):
    """
    Around the branch-point centroid, find pairs of endpoints whose
    exit angles are roughly opposite (≈180°).  Returns list of (ep_a, ep_b).
    """
    if not branch_points or len(endpoints) < 2:
        return []

    centroid = np.array(branch_points).mean(axis=0)
    nearby   = [
        {"ep": ep,
         "dist": np.linalg.norm(np.array(ep) - centroid),
         "angle": _endpoint_exit_angle(ep, graph)}
        for ep in endpoints
        if np.linalg.norm(np.array(ep) - centroid) <= CLUSTER_RADIUS_PX
    ]
    if len(nearby) < 2:
        return []

    scored = []
    for a, b in combinations(nearby, 2):
        diff      = abs(a["angle"] - b["angle"])
        diff      = min(diff, 360 - diff)
        deviation = abs(diff - 180.0)
        if deviation <= ANGLE_TOLERANCE_DEG:
            scored.append((deviation, a["ep"], b["ep"]))

    scored.sort(key=lambda x: x[0])
    used, pairs = set(), []
    for _, ep_a, ep_b in scored:
        if ep_a not in used and ep_b not in used:
            pairs.append((ep_a, ep_b))
            used.update({ep_a, ep_b})

    log.info(f"  Endpoint pairing: {len(nearby)} nearby → {len(pairs)} pair(s)")
    return pairs


def _draw_overlap_debug(overlay, endpoints, branch_points, pairs, centroid=None):
    """Annotate overlay with branch points, endpoint markers, and pair lines."""
    for bp in branch_points:
        cv2.circle(overlay, (bp[1], bp[0]), 3, (0, 255, 255), -1)
    for ep in endpoints:
        cv2.circle(overlay, (ep[1], ep[0]), 5, (0, 220, 220), 1)
    if centroid is not None:
        cv2.drawMarker(overlay,
                       (int(centroid[1]), int(centroid[0])),
                       (255, 255, 255), cv2.MARKER_CROSS, 12, 2)
    colors = [(255,0,255),(255,128,0),(0,255,128),(128,0,255),(255,255,0)]
    for i, (ep_a, ep_b) in enumerate(pairs):
        c = colors[i % len(colors)]
        cv2.line(overlay, (ep_a[1], ep_a[0]), (ep_b[1], ep_b[0]), c, 1)
        cv2.circle(overlay, (ep_a[1], ep_a[0]), 6, c, 2)
        cv2.circle(overlay, (ep_b[1], ep_b[0]), 6, c, 2)

# ============================================================
# ⓮  MEASURE A SINGLE STRAW (one pair of endpoints)
# ============================================================

def _measure_pair(ep_a, ep_b, graph, straw_id, pixel_to_mm,
                  overlay, source, yolo_conf, is_reconstructed=False):
    """
    Trace the Dijkstra path between ep_a and ep_b, convert to mm, annotate
    the overlay, and return a result dict.  Returns None on failure.
    """
    path_len, path_nodes = dijkstra_path(graph, ep_a, ep_b)

    if path_len == float("inf") or path_len < MIN_PATH_LENGTH_PX:
        return None

    euclid = np.linalg.norm(np.array(ep_a) - np.array(ep_b))
    if euclid == 0:
        return None

    ratio     = path_len / euclid
    shape     = "Curved" if ratio > CURVE_THRESHOLD else "Straight"
    length_mm = path_len * pixel_to_mm

    if   is_reconstructed:    confidence = "RECONSTRUCTED"
    elif source == "yolo":
        confidence = ("HIGH" if yolo_conf >= 0.7 else
                      "MEDIUM" if yolo_conf >= 0.4 else "LOW")
    else:
        confidence = "HIGH"

    log.info(f"  Straw {straw_id}: {length_mm:.2f} mm | {shape} | {confidence}")

    # Draw path on overlay
    path_color = (0, 140, 255) if is_reconstructed else (0, 200, 255)
    for node in path_nodes:
        overlay[node[0], node[1]] = path_color
    ep_color = (255, 80, 0) if is_reconstructed else (0, 255, 0)
    cv2.circle(overlay, (ep_a[1], ep_a[0]), 5, ep_color, -1)
    cv2.circle(overlay, (ep_b[1], ep_b[0]), 5, ep_color, -1)
    cx = int((ep_a[1] + ep_b[1]) / 2)
    cy = int((ep_a[0] + ep_b[0]) / 2) - 8
    tag = "R" if is_reconstructed else ""
    cv2.putText(overlay,
                f"#{straw_id}{tag} {length_mm:.1f}mm [{confidence}]",
                (cx, cy), cv2.FONT_HERSHEY_SIMPLEX,
                0.38, (255, 255, 255), 1, cv2.LINE_AA)

    return {
        "Straw_ID":             straw_id,
        "Length_mm":            round(length_mm, 3),
        "Shape":                shape,
        "Overlap_Detected":     is_reconstructed,
        "Reconstructed":        is_reconstructed,
        "Confidence":           confidence,
        "Detection_Source":     source,
        "YOLO_Conf":            round(yolo_conf, 3) if source == "yolo" else "N/A",
        "Skeleton_Pixels":      len(path_nodes),
        "Path_Euclidean_Ratio": round(ratio, 4),
        "Pixel_to_mm_Used":     pixel_to_mm,
    }

# ============================================================
# ⓯  MEASURE ONE MASK  (dispatcher: BFS or Dijkstra)
# ============================================================

def measure_mask(mask_entry, base_straw_id, pixel_to_mm, overlay):
    """
    Full measurement pipeline for a single mask:
      1. Clean + remove tiny blobs
      2. Skeletonize
      3. Classify skeleton points
      4. If branch points exist → Dijkstra overlap reconstruction
      5. Otherwise             → BFS longest path (fast)
      6. Apply length guard-rails
    Returns a list of result dicts (may be >1 if crossing straws).
    """
    label     = f"straw {base_straw_id}"
    m         = mask_entry["mask"]
    yolo_conf = mask_entry["confidence"]
    source    = mask_entry["source"]

    # --- 1. Clean mask ---
    raw_area = int(np.sum(m > 0))
    m = clean_mask(m)
    m = merge_small_components(m)

    post_area = int(np.sum(m > 0))
    if post_area == 0:
        log.warning(f"  [{label}] REJECTED — empty after cleaning "
                    f"(raw area={raw_area}px)")
        return []

    # Slight dilation to bridge small skeleton gaps
    m = cv2.dilate(m, np.ones((2, 2), np.uint8), iterations=1)

    # --- 2. Skeletonize ---
    skeleton = skeletonize(m // 255)
    points   = np.column_stack(np.where(skeleton > 0))

    if len(points) < 3:
        log.warning(f"  [{label}] REJECTED — skeleton too sparse "
                    f"({len(points)} pts, post-clean area={post_area}px)")
        return []

    # --- 3. Classify skeleton points ---
    endpoints, body_pts, branch_pts = classify_skeleton_points(points)

    # --- 4. OVERLAP PATH: Dijkstra + endpoint pairing ---
    if branch_pts and len(endpoints) >= 2:
        log.info(f"  [{label}] Overlap detected "
                 f"({len(branch_pts)} branch pts, {len(endpoints)} endpoints)")

        graph    = _build_graph(points)
        pairs    = cluster_and_pair_endpoints(endpoints, branch_pts, graph)
        centroid = np.array(branch_pts).mean(axis=0) if branch_pts else None
        _draw_overlap_debug(overlay, endpoints, branch_pts, pairs, centroid)

        if pairs:
            results = []
            for sub_id, (ep_a, ep_b) in enumerate(pairs):
                straw_id = f"{base_straw_id}.{sub_id}"
                res = _measure_pair(ep_a, ep_b, graph, straw_id,
                                    pixel_to_mm, overlay, source,
                                    yolo_conf, is_reconstructed=True)
                if res is None:
                    continue
                # Guard-rails
                lmm = res["Length_mm"]
                if lmm < MIN_LENGTH_MM:
                    log.info(f"  [{straw_id}] REJECTED — too short ({lmm:.1f}mm)")
                    continue
                if lmm > MAX_LENGTH_MM:
                    log.info(f"  [{straw_id}] REJECTED — too long ({lmm:.1f}mm)")
                    continue
                results.append(res)
            if results:
                return results
            log.warning(f"  [{label}] Pairing produced no valid straws — "
                        f"falling back to BFS.")

    # --- 5. NO OVERLAP PATH: BFS longest path ---
    length_px = longest_path_bfs(points)
    length_mm = length_px * pixel_to_mm

    log.info(f"  [{label}] BFS path={length_px:.1f}px → {length_mm:.1f}mm "
             f"(skeleton pts={len(points)})")

    if length_px < MIN_PATH_LENGTH_PX:
        log.info(f"  [{label}] REJECTED — path too short ({length_px:.1f}px)")
        return []
    if length_mm < MIN_LENGTH_MM:
        log.info(f"  [{label}] REJECTED — too short ({length_mm:.1f}mm)")
        return []
    if length_mm > MAX_LENGTH_MM:
        log.info(f"  [{label}] REJECTED — too long ({length_mm:.1f}mm)")
        return []

    # Annotate overlay
    pts_arr = np.array(list(map(list, [tuple(p) for p in points])))
    dm      = cdist(pts_arr, pts_arr)
    idx     = np.unravel_index(np.argmax(dm), dm.shape)
    ep_a    = tuple(pts_arr[idx[0]])
    ep_b    = tuple(pts_arr[idx[1]])

    if source == "yolo":
        conf_label = ("HIGH" if yolo_conf >= 0.7 else
                      "MEDIUM" if yolo_conf >= 0.4 else "LOW")
    else:
        conf_label = "HIGH"

    cv2.circle(overlay, (int(ep_a[1]), int(ep_a[0])), 5, (0, 255, 0), -1)
    cv2.circle(overlay, (int(ep_b[1]), int(ep_b[0])), 5, (0, 255, 0), -1)
    cx = int((ep_a[1] + ep_b[1]) / 2)
    cy = int((ep_a[0] + ep_b[0]) / 2) - 8
    cv2.putText(overlay,
                f"#{base_straw_id} {length_mm:.1f}mm [{conf_label}]",
                (cx, cy), cv2.FONT_HERSHEY_SIMPLEX,
                0.38, (255, 255, 255), 1, cv2.LINE_AA)

    return [{
        "Straw_ID":             base_straw_id,
        "Length_mm":            round(length_mm, 3),
        "Shape":                "Straight",
        "Overlap_Detected":     False,
        "Reconstructed":        False,
        "Confidence":           conf_label,
        "Detection_Source":     source,
        "YOLO_Conf":            round(yolo_conf, 3) if source == "yolo" else "N/A",
        "Skeleton_Pixels":      len(points),
        "Path_Euclidean_Ratio": None,
        "Pixel_to_mm_Used":     pixel_to_mm,
        "Endpoint_Method":      "bfs",
    }]

# ============================================================
# ⓰  CALIBRATION — QUARTER COIN
# ============================================================

def calibrate_from_quarter(image, yolo_model):
    """
    Detect a quarter coin in image using yolo_model.
    Returns (pixel_to_mm, method_str) or (None, reason_str).
    Tries ellipse fit on the segmentation mask first; falls back to bbox.
    """
    if yolo_model is None:
        return None, "no YOLO model"

    h_img, w_img = image.shape[:2]
    results      = yolo_model(image, conf=YOLO_CONF, verbose=False)[0]
    names        = yolo_model.names
    detections   = []

    for idx, box in enumerate(results.boxes):
        cls_name = names.get(int(box.cls[0]), "").lower().strip()
        if cls_name != QUARTER_CLASS_NAME.lower().strip():
            continue

        conf           = float(box.conf[0])
        x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
        pixel_diameter = None
        method         = None

        # Method 1: ellipse on segmentation mask
        if results.masks is not None and idx < len(results.masks.data):
            mask_raw = results.masks.data[idx].cpu().numpy()
            mask_u8  = (mask_raw * 255).astype(np.uint8)
            mask_u8  = cv2.resize(mask_u8, (w_img, h_img),
                                  interpolation=cv2.INTER_NEAREST)
            _, mask_bin = cv2.threshold(mask_u8, 127, 255, cv2.THRESH_BINARY)
            contours, _ = cv2.findContours(mask_bin, cv2.RETR_EXTERNAL,
                                           cv2.CHAIN_APPROX_SIMPLE)
            if contours:
                largest = max(contours, key=cv2.contourArea)
                if len(largest) >= 5:
                    try:
                        ell            = cv2.fitEllipse(largest)
                        pixel_diameter = (max(ell[1]) + min(ell[1])) / 2.0
                        method         = "ellipse"
                        log.info(f"  Quarter ellipse: d={pixel_diameter:.1f}px")
                    except cv2.error as exc:
                        log.warning(f"  fitEllipse failed: {exc}")

        # Method 2: bounding-box fallback
        if pixel_diameter is None:
            pixel_diameter = ((x2 - x1) + (y2 - y1)) / 2.0
            method         = "bbox_fallback"
            log.info(f"  Quarter bbox: d={pixel_diameter:.1f}px")

        if pixel_diameter and pixel_diameter > 0:
            detections.append({
                "pixel_diameter": pixel_diameter,
                "method":         method,
                "conf":           conf,
            })

    if not detections:
        return None, "no quarter detected"

    best  = max(detections, key=lambda d: d["conf"])
    px2mm = QUARTER_DIAMETER_MM / best["pixel_diameter"]
    log.info(f"  Calibration: {QUARTER_DIAMETER_MM}mm / "
             f"{best['pixel_diameter']:.2f}px = {px2mm:.5f} mm/px "
             f"[{best['method']} conf={best['conf']:.2f}]")
    return px2mm, best["method"]


def get_calibration_for_image(image, image_path, yolo_model, cached_ratio):
    """
    Return (pixel_to_mm, updated_cached_ratio).
    Respects CALIBRATION_PER_IMAGE and falls back to cache or PIXEL_TO_MM.
    """
    fname = os.path.basename(image_path)

    if CALIBRATION_MODE == "manual":
        return _manual_calibration(), cached_ratio

    if not CALIBRATION_PER_IMAGE and cached_ratio is not None:
        log.info(f"  [{fname}] Reusing cached calibration: {cached_ratio:.5f} mm/px")
        return cached_ratio, cached_ratio

    ratio, info = calibrate_from_quarter(image, yolo_model)

    if ratio is None:
        if cached_ratio is not None:
            log.warning(f"  [{fname}] No quarter — reusing cached ratio.")
            return cached_ratio, cached_ratio
        log.warning(f"  [{fname}] No quarter and no cache — "
                    f"using default {PIXEL_TO_MM} mm/px.")
        print(f"\n  WARNING: '{fname}' — no quarter found.  "
              f"Using default {PIXEL_TO_MM} mm/px.  "
              f"(Set CALIBRATION_MODE='manual' or include a quarter.)\n")
        return PIXEL_TO_MM, cached_ratio

    log.info(f"  [{fname}] Auto-calibration: {ratio:.5f} mm/px [{info}]")
    return ratio, ratio


def _manual_calibration():
    print(f"\n--- CALIBRATION ---\nDefault: {PIXEL_TO_MM} mm/px")
    raw = input("Enter mm per pixel (or Enter for default): ").strip()
    if raw:
        try:
            return float(raw)
        except ValueError:
            log.warning("Invalid input. Using default.")
    return PIXEL_TO_MM

# ============================================================
# ⓱  PROCESS A SINGLE IMAGE
# ============================================================

def process_image(image_path, pixel_to_mm, mode,
                  hsv_lower=None, hsv_upper=None, yolo_model=None):
    """
    Load, detect, measure, annotate.
    Returns (list_of_result_dicts, overlay_image).
    """
    log.info(f"\n{'='*60}")
    log.info(f"Processing [{mode.upper()}]: {os.path.basename(image_path)}")

    image = cv2.imread(image_path)
    if image is None:
        log.error(f"Failed to load: {image_path}")
        return [], None
    image = cv2.resize(image, (800, 600))
    debug_step("Original", image)
    overlay = image.copy()

    # --- Get masks ---
    yolo_results = None
    if mode == "yolo" and yolo_model is not None:
        mask_entries, yolo_results = get_masks_yolo(image, yolo_model)
        if not mask_entries:
            log.warning("  YOLO returned 0 masks — falling back to HSV.")
            mask_entries = get_masks_hsv(
                image,
                hsv_lower or STRAW_HSV_LOWER,
                hsv_upper or STRAW_HSV_UPPER,
            )
    else:
        if mode == "yolo":
            log.warning("  YOLO model unavailable — falling back to HSV.")
        mask_entries = get_masks_hsv(
            image,
            hsv_lower or STRAW_HSV_LOWER,
            hsv_upper or STRAW_HSV_UPPER,
        )

    log.info(f"  Masks to process: {len(mask_entries)}")

    # --- Measure each mask ---
    all_results = []
    for i, entry in enumerate(mask_entries):
        res_list = measure_mask(entry, i + 1, pixel_to_mm, overlay)
        for res in res_list:
            res["Image"]       = os.path.basename(image_path)
            res["Pixel_to_mm"] = round(pixel_to_mm, 6)
            all_results.append(res)

    log.info(f"  Done: {len(all_results)} straw(s) from {len(mask_entries)} mask(s).")

    # HUD
    calib_note = f"px→mm:{pixel_to_mm:.5f} ({'auto' if CALIBRATION_MODE == 'auto' else 'manual'})"
    cv2.putText(overlay,
                f"Mode:{mode.upper()}  Straws:{len(all_results)}  "
                f"Masks:{len(mask_entries)}  {calib_note}",
                (10, 580), cv2.FONT_HERSHEY_SIMPLEX,
                0.38, (200, 200, 200), 1, cv2.LINE_AA)

    if not _SUPPRESS_POPUPS:
        debug_step("Final Overlay", overlay, wait=800)

    return all_results, overlay

# ============================================================
# ⓲  SUMMARY PRINT
# ============================================================

def print_summary(df):
    print("\n" + "=" * 60)
    print("BATCH SUMMARY")
    print("=" * 60)
    print(f"Total straws measured   :  {len(df)}")
    if len(df) > 0:
        print(f"Average length (mm)     :  {df['Length_mm'].mean():.2f}")
        print(f"Min / Max (mm)          :  "
              f"{df['Length_mm'].min():.2f} / {df['Length_mm'].max():.2f}")
        print(f"Straight / Curved       :  "
              f"{(df['Shape']=='Straight').sum()} / "
              f"{(df['Shape']=='Curved').sum()}")
        print(f"Overlap detected        :  {df['Overlap_Detected'].sum()}")
        if "Reconstructed" in df.columns:
            print(f"Reconstructed (crossed) :  {df['Reconstructed'].sum()}")
        for conf in ("HIGH", "MEDIUM", "LOW", "RECONSTRUCTED"):
            if (df["Confidence"] == conf).any():
                print(f"{conf:24s}:  {(df['Confidence']==conf).sum()}")
        if "Pixel_to_mm" in df.columns:
            vals = df["Pixel_to_mm"].unique()
            if len(vals) == 1:
                print(f"Calibration (mm/px)     :  {vals[0]:.6f}")
            else:
                print(f"Calibration (mm/px)     :  "
                      f"min={vals.min():.6f}  max={vals.max():.6f} "
                      f"[{len(vals)} values]")
        if "Detection_Source" in df.columns:
            print(f"YOLO / HSV detections   :  "
                  f"{(df['Detection_Source']=='yolo').sum()} / "
                  f"{(df['Detection_Source']=='hsv').sum()}")
    print("=" * 60 + "\n")

# ============================================================
# ⓳  EXCEL EXPORT WITH EMBEDDED OVERLAY IMAGES
# ============================================================

def save_excel_with_images(df, overlay_paths, filename):
    """
    Sheet 1 — Results: formatted data table.
    Sheet 2 — Overlays: each processed image embedded, labelled.
    """
    wb = Workbook()

    # ── Sheet 1: Results ──────────────────────────────────────
    ws   = wb.active
    ws.title = "Results"

    hdr_fill   = PatternFill("solid", fgColor="2E4057")
    hdr_font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    row_font   = Font(name="Arial", size=10)
    alt_fill   = PatternFill("solid", fgColor="EAF4FB")
    bd         = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=bd, right=bd, top=bd, bottom=bd)

    col_order = [
        "Image", "Straw_ID", "Length_mm", "Shape",
        "Overlap_Detected", "Reconstructed", "Confidence",
        "Detection_Source", "YOLO_Conf", "Endpoint_Method",
        "Skeleton_Pixels", "Path_Euclidean_Ratio",
        "Pixel_to_mm_Used", "Pixel_to_mm",
    ]
    cols = [c for c in col_order if c in df.columns]

    for ci, name in enumerate(cols, 1):
        cell           = ws.cell(row=1, column=ci, value=name)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = cell_border

    for ri, (_, row) in enumerate(df[cols].iterrows(), 2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, col in enumerate(cols, 1):
            cell           = ws.cell(row=ri, column=ci, value=row[col])
            cell.font      = row_font
            cell.border    = cell_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                cell.fill = fill

    for ci, col in enumerate(cols, 1):
        max_len = max(
            len(str(col)),
            max((len(str(df[col].iloc[i])) for i in range(len(df))), default=0),
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 30)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # ── Sheet 2: Overlays ─────────────────────────────────────
    ws2 = wb.create_sheet(title="Overlays")
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 100

    for col, label in (("A1", "Image"), ("B1", "Overlay")):
        ws2[col] = label
        ws2[col].font      = hdr_font
        ws2[col].fill      = hdr_fill
        ws2[col].alignment = Alignment(horizontal="center", vertical="center")

    img_row        = 2
    IMG_HEIGHT_PX  = 300
    ROW_HEIGHT_PT  = 225

    for img_path, fname, n_straws in overlay_paths:
        if not os.path.exists(img_path):
            log.warning(f"  Overlay not found: {img_path}")
            continue

        cell           = ws2.cell(row=img_row, column=1,
                                  value=f"{fname}\n{n_straws} straw(s)")
        cell.font      = Font(name="Arial", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        try:
            xl_img        = XLImage(img_path)
            scale         = IMG_HEIGHT_PX / xl_img.height if xl_img.height > 0 else 1.0
            xl_img.height = IMG_HEIGHT_PX
            xl_img.width  = int(xl_img.width * scale)
            xl_img.anchor = f"{get_column_letter(2)}{img_row}"
            ws2.add_image(xl_img)
            ws2.row_dimensions[img_row].height = ROW_HEIGHT_PT
            log.info(f"  Embedded overlay: {img_path}")
        except Exception as exc:
            ws2.cell(row=img_row, column=2,
                     value=f"[Could not embed: {exc}]")
        img_row += 1

    # ── Save (handle locked file) ─────────────────────────────
    base_path = os.path.abspath(filename)
    try:
        wb.save(base_path)
        return base_path
    except PermissionError:
        ts        = time.strftime("%Y%m%d_%H%M%S")
        name, ext = os.path.splitext(base_path)
        new_path  = f"{name}_{ts}{ext}"
        log.warning(f"'{base_path}' locked — saving to '{new_path}'")
        print(f"\n  WARNING: File open in Excel. Saving to "
              f"{os.path.basename(new_path)}\n")
        wb.save(new_path)
        return new_path

# ============================================================
# ⓴  MAIN
# ============================================================

if __name__ == "__main__":

    t0 = time.time()
    log.info("=== Straw Measurement Tool Started ===")
    log.info(f"Detection mode      : {DETECTION_MODE.upper()}")
    log.info(f"Calibration mode    : {CALIBRATION_MODE.upper()}")
    log.info(f"Calibration per img : {CALIBRATION_PER_IMAGE}")
    log.info(f"YOLO model          : {YOLO_MODEL_PATH}")
    log.info(f"YOLO conf threshold : {YOLO_CONF}")
    log.info(f"Length range (mm)   : {MIN_LENGTH_MM} – {MAX_LENGTH_MM}")
    log.info(f"Angle tolerance     : {ANGLE_TOLERANCE_DEG}°  "
             f"Cluster radius: {CLUSTER_RADIUS_PX}px")

    # ── File selection ────────────────────────────────────────
    files = select_files()
    if not files:
        log.error("No files selected. Exiting.")
        sys.exit(0)

    # ── Mode setup ────────────────────────────────────────────
    mode       = DETECTION_MODE.lower()
    yolo_model = None
    hsv_lower  = hsv_upper = None

    if mode == "yolo":
        yolo_model = load_yolo_model(YOLO_MODEL_PATH)
        if yolo_model is None:
            log.warning("YOLO unavailable — falling back to HSV mode.")
            mode = "hsv"

    # Suppress all cv2 popup windows in YOLO mode
    if mode == "yolo":
        _SUPPRESS_POPUPS = True
        log.info("YOLO mode: popup windows suppressed. "
                 "Overlays will be embedded in Excel.")

    # Load YOLO for calibration even if detection is HSV
    calib_model = yolo_model or (
        load_yolo_model(YOLO_MODEL_PATH)
        if CALIBRATION_MODE == "auto" else None
    )

    # HSV tuning (only in HSV mode)
    if mode == "hsv":
        hsv_lower, hsv_upper = get_hsv(files[0])

    # Manual global calibration
    global_pixel_to_mm = (
        _manual_calibration() if CALIBRATION_MODE == "manual" else None
    )

    # ── Batch processing ──────────────────────────────────────
    cached_ratio  = None
    all_results   = []
    skipped       = []
    overlay_paths = []   # (saved_path, filename, straw_count)

    for idx, f in enumerate(files):
        log.info(f"\nImage {idx+1}/{len(files)}: {f}")

        raw_image = cv2.imread(f)
        if raw_image is None:
            log.error(f"Failed to load: {f}")
            skipped.append(f)
            continue
        raw_image = cv2.resize(raw_image, (800, 600))

        if CALIBRATION_MODE == "manual":
            pixel_to_mm = global_pixel_to_mm
        else:
            pixel_to_mm, cached_ratio = get_calibration_for_image(
                raw_image, f, calib_model, cached_ratio
            )

        if pixel_to_mm is None:
            log.warning(f"  No calibration for {os.path.basename(f)} — skipping.")
            skipped.append(f)
            continue

        res_list, overlay = process_image(
            f, pixel_to_mm, mode,
            hsv_lower=hsv_lower,
            hsv_upper=hsv_upper,
            yolo_model=yolo_model,
        )
        all_results.extend(res_list)

        if overlay is not None:
            base    = os.path.splitext(os.path.basename(f))[0]
            out_dir = os.path.dirname(os.path.abspath(f))
            out_img = os.path.join(out_dir, f"overlay_{base}.jpg")
            cv2.imwrite(out_img, overlay)
            log.info(f"  Overlay saved: {out_img}")
            overlay_paths.append((out_img, os.path.basename(f), len(res_list)))

            if not _SUPPRESS_POPUPS:
                cv2.imshow(
                    f"Result [{idx+1}/{len(files)}]: {os.path.basename(f)}",
                    overlay,
                )
                cv2.waitKey(800)

    if not _SUPPRESS_POPUPS:
        cv2.destroyAllWindows()

    # ── Skipped files report ──────────────────────────────────
    if skipped:
        print(f"\n  {len(skipped)} image(s) skipped:")
        for s in skipped:
            print(f"    {s}")

    # ── Save results ──────────────────────────────────────────
    df = pd.DataFrame(all_results)

    if len(df) == 0:
        log.warning("No straws detected across all images.")
        print("\nNo straws detected. Check YOLO model path or HSV settings.")
        sys.exit(0)

    col_order = [
        "Image", "Straw_ID", "Length_mm", "Shape",
        "Overlap_Detected", "Reconstructed", "Confidence",
        "Detection_Source", "YOLO_Conf", "Endpoint_Method",
        "Skeleton_Pixels", "Path_Euclidean_Ratio",
        "Pixel_to_mm_Used", "Pixel_to_mm",
    ]
    df = df[[c for c in col_order if c in df.columns]]

    xlsx_path = os.path.abspath("straw_results.xlsx")
    if mode == "yolo" and overlay_paths:
        xlsx_path = save_excel_with_images(df, overlay_paths, xlsx_path)
        print(f"\nOverlay images embedded in Excel on the 'Overlays' sheet.")
    else:
        try:
            df.to_excel(xlsx_path, index=False)
        except PermissionError:
            ts        = time.strftime("%Y%m%d_%H%M%S")
            name, ext = os.path.splitext(xlsx_path)
            xlsx_path = f"{name}_{ts}{ext}"
            df.to_excel(xlsx_path, index=False)

    csv_path = os.path.abspath("straw_results.csv")
    df.to_csv(csv_path, index=False)

    print_summary(df)
    print(f"Results saved to:\n  Excel : {xlsx_path}\n  CSV   : {csv_path}")
    log.info(f"=== Done in {time.time() - t0:.2f}s ===")
    print(f"\nDebug log: {os.path.abspath(LOG_FILE)}")